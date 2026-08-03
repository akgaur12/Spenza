"""File parsing for the import module: turns an uploaded CSV/XLSX file into
a stream of raw, untyped row values keyed by the four required logical
columns. No business validation happens here — see `validators.py` and
`import_service.py` for parsing/resolving those raw values.

Both parsers are generators so a caller (see `ImportService._parse_rows`)
can enforce a row-count limit while consuming them, without first buffering
an arbitrarily large file fully into a list.
"""

import csv
import io
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from src.modules.import_export.exceptions import (
    ImportFileEmptyError,
    ImportFileUnreadableError,
    ImportMissingColumnsError,
)

REQUIRED_COLUMNS = ("date", "category", "description", "amount")


@dataclass(frozen=True, slots=True)
class RawImportRow:
    """One data row exactly as read from the file — no parsing/validation
    applied yet. `date` and `amount` may already be native Python
    `date`/`datetime`/`int`/`float` values for XLSX (Excel-typed cells)
    rather than strings.
    """

    row_number: int
    date: Any
    category: Any
    description: Any
    amount: Any


def _normalize_header(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def _column_indices(header_row: list[object]) -> dict[str, int]:
    normalized = [_normalize_header(v) for v in header_row]
    indices = {name: normalized.index(name) for name in REQUIRED_COLUMNS if name in normalized}
    missing = [name for name in REQUIRED_COLUMNS if name not in indices]
    if missing:
        raise ImportMissingColumnsError(
            message=f"Missing required column(s): {', '.join(missing)}",
            details={"missing_columns": missing},
        )
    return indices


def _cell(row: list[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


def parse_csv_rows(content: bytes) -> Iterator[RawImportRow]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportFileUnreadableError(message="File is not valid UTF-8 text") from exc

    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration as exc:
        raise ImportFileEmptyError() from exc

    indices = _column_indices(list(header))
    row_number = 1
    has_rows = False
    for raw_row in reader:
        row_number += 1
        if not any(cell.strip() for cell in raw_row):
            continue
        has_rows = True
        yield RawImportRow(
            row_number=row_number,
            date=_cell(raw_row, indices["date"]),
            category=_cell(raw_row, indices["category"]),
            description=_cell(raw_row, indices["description"]),
            amount=_cell(raw_row, indices["amount"]),
        )
    if not has_rows:
        raise ImportFileEmptyError()


def parse_xlsx_rows(content: bytes) -> Iterator[RawImportRow]:
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
    except (InvalidFileException, KeyError, OSError, zipfile.BadZipFile) as exc:
        raise ImportFileUnreadableError(
            message="File could not be read as an XLSX workbook"
        ) from exc

    try:
        sheet = workbook.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise ImportFileEmptyError() from exc

        indices = _column_indices(list(header))
        row_number = 1
        has_rows = False
        for raw_row in rows_iter:
            row_number += 1
            if raw_row is None or all(v is None for v in raw_row):
                continue
            has_rows = True
            yield RawImportRow(
                row_number=row_number,
                date=_cell(list(raw_row), indices["date"]),
                category=_cell(list(raw_row), indices["category"]),
                description=_cell(list(raw_row), indices["description"]),
                amount=_cell(list(raw_row), indices["amount"]),
            )
        if not has_rows:
            raise ImportFileEmptyError()
    finally:
        workbook.close()
