"""Pure functions turning a list of `ExportRow` into CSV/XLSX/PDF bytes.

No DB access, no FastAPI concerns — kept separate from `export_service.py`'s
expense retrieval so each format's layout can be tested in isolation.

Dates are rendered `DD-MMM-YYYY` and weekdays as the fixed 3-letter form
(`Mon`..`Sun`) via lookup tables rather than `strftime("%b"/"%a")`, since
both are locale-dependent and would silently render differently on a server
not configured for an English locale.
"""

import io
from csv import writer as csv_writer
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from src.core.timezone import APP_TIMEZONE

_MONTH_ABBR = (
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
)
_WEEKDAY_ABBR = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")

# reportlab's built-in fonts can't reliably render "₹" (U+20B9) without
# shipping a Unicode-capable TTF, which this project doesn't otherwise need —
# per the spec, fall back to a safe plain-text currency label instead of
# risking a broken glyph.
CURRENCY_FALLBACK = "INR"

HEADERS = ("Date", "Day", "Category", "Description", "Amount")


@dataclass(frozen=True, slots=True)
class ExportRow:
    spent_date: date
    category_name: str
    description: str
    amount: Decimal


def format_export_date(value: date) -> str:
    return f"{value.day:02d}-{_MONTH_ABBR[value.month - 1]}-{value.year:04d}"


def weekday_abbr(value: date) -> str:
    return _WEEKDAY_ABBR[value.weekday()]


def build_csv_export(rows: list[ExportRow]) -> bytes:
    buffer = io.StringIO()
    writer = csv_writer(buffer)
    writer.writerow(HEADERS)
    for row in rows:
        writer.writerow(
            [
                format_export_date(row.spent_date),
                weekday_abbr(row.spent_date),
                row.category_name,
                row.description,
                f"{row.amount:.2f}",
            ]
        )
    return buffer.getvalue().encode("utf-8-sig")


def build_xlsx_export(rows: list[ExportRow]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None  # a freshly created Workbook always has an active sheet
    sheet.title = "Expenses"

    sheet.append(list(HEADERS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for row in rows:
        sheet.append(
            [
                format_export_date(row.spent_date),
                weekday_abbr(row.spent_date),
                row.category_name,
                row.description,
                float(row.amount),
            ]
        )
        sheet.cell(row=sheet.max_row, column=5).number_format = "0.00"

    for index, width in enumerate((14, 6, 18, 40, 14), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    if rows:
        sheet.auto_filter.ref = f"A1:E{sheet.max_row}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_pdf_export(
    rows: list[ExportRow], *, start_date: date | None, end_date: date | None
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        title="Expense Report",
        # Uncompressed content streams keep the rendered text inspectable as
        # plain bytes — this project has no PDF-parsing dependency, and the
        # test suite deliberately avoids fragile pixel-perfect PDF assertions
        # in favor of checking that expected text made it into the output.
        pageCompression=0,
    )
    styles = getSampleStyleSheet()

    total_spending = sum((r.amount for r in rows), Decimal("0.00"))
    generated_at = datetime.now(UTC).astimezone(APP_TIMEZONE)

    elements: list[object] = [Paragraph("Expense Report", styles["Title"]), Spacer(1, 6)]
    if start_date is not None and end_date is not None:
        elements.append(
            Paragraph(
                f"Period: {format_export_date(start_date)} to {format_export_date(end_date)}",
                styles["Normal"],
            )
        )
    elements.append(
        Paragraph(f"Generated At: {generated_at.strftime('%d-%b-%Y %H:%M')}", styles["Normal"])
    )
    elements.append(Paragraph(f"Total Expenses: {len(rows)}", styles["Normal"]))
    elements.append(
        Paragraph(f"Total Spending: {CURRENCY_FALLBACK} {total_spending:,.2f}", styles["Normal"])
    )
    elements.append(Spacer(1, 12))

    if not rows:
        elements.append(Paragraph("No expenses found for the selected period.", styles["Normal"]))
    else:
        table_data: list[list[str]] = [list(HEADERS)]
        for row in rows:
            table_data.append(
                [
                    format_export_date(row.spent_date),
                    weekday_abbr(row.spent_date),
                    row.category_name,
                    row.description,
                    f"{CURRENCY_FALLBACK} {row.amount:,.2f}",
                ]
            )
        table = Table(table_data, repeatRows=1, colWidths=[70, 35, 90, 180, 80])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d3748")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f7fafc")],
                    ),
                ]
            )
        )
        elements.append(table)

    doc.build(elements)
    return buffer.getvalue()
