"""Integration tests for GET /api/v1/export/expenses?format=xlsx."""

import io

import openpyxl
from httpx import AsyncClient

from tests.conftest import RecordingEmailBackend
from tests.import_export.helpers import (
    category_id_by_name,
    create_expense,
    login_user_a,
    login_user_b,
)


def _load_sheet(content: bytes) -> tuple[object, list[tuple[object, ...]]]:
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active
    assert sheet is not None
    rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    return sheet, rows


async def test_export_xlsx_requires_authentication(client: AsyncClient) -> None:
    response = await client.get("/api/v1/export/expenses", params={"format": "xlsx"})
    assert response.status_code == 401


async def test_export_xlsx_valid_workbook_and_worksheet_name(
    client: AsyncClient, email_backend: RecordingEmailBackend
) -> None:
    await login_user_a(client, email_backend)
    response = await client.get("/api/v1/export/expenses", params={"format": "xlsx"})
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert workbook.active is not None
    assert workbook.active.title == "Expenses"


async def test_export_xlsx_header_row(
    client: AsyncClient, email_backend: RecordingEmailBackend
) -> None:
    await login_user_a(client, email_backend)
    response = await client.get("/api/v1/export/expenses", params={"format": "xlsx"})
    _, rows = _load_sheet(response.content)
    assert rows[0] == ("Date", "Day", "Category", "Description", "Amount")


async def test_export_xlsx_values_and_amount_is_numeric(
    client: AsyncClient, email_backend: RecordingEmailBackend
) -> None:
    await login_user_a(client, email_backend)
    food_id = await category_id_by_name(client, "Food")
    await create_expense(
        client,
        category_id=food_id,
        description="Cake",
        amount="278.00",
        spent_at="2025-01-01T00:00:00+05:30",
    )
    response = await client.get("/api/v1/export/expenses", params={"format": "xlsx"})
    _, rows = _load_sheet(response.content)
    date_value, day_value, category_value, description_value, amount_value = rows[1]
    assert date_value == "01-Jan-2025"
    assert day_value == "Wed"
    assert category_value == "Food"
    assert description_value == "Cake"
    assert isinstance(amount_value, int | float)
    assert float(amount_value) == 278.0


async def test_export_xlsx_filters_work(
    client: AsyncClient, email_backend: RecordingEmailBackend
) -> None:
    await login_user_a(client, email_backend)
    food_id = await category_id_by_name(client, "Food")
    transport_id = await category_id_by_name(client, "Transport")
    await create_expense(
        client, category_id=food_id, spent_at="2025-01-01T00:00:00+05:30", description="Cake"
    )
    await create_expense(
        client, category_id=transport_id, spent_at="2025-01-02T00:00:00+05:30", description="Petrol"
    )
    response = await client.get(
        "/api/v1/export/expenses", params={"format": "xlsx", "category_id": food_id}
    )
    _, rows = _load_sheet(response.content)
    descriptions = [row[3] for row in rows[1:]]
    assert descriptions == ["Cake"]


async def test_export_xlsx_user_isolation(
    client: AsyncClient, email_backend: RecordingEmailBackend
) -> None:
    await login_user_b(client, email_backend)
    food_id_b = await category_id_by_name(client, "Food")
    await create_expense(
        client, category_id=food_id_b, spent_at="2025-01-01T00:00:00+05:30", description="UserB"
    )

    await login_user_a(client, email_backend)
    response = await client.get("/api/v1/export/expenses", params={"format": "xlsx"})
    _, rows = _load_sheet(response.content)
    assert len(rows) == 1  # header only — user A has no expenses


async def test_export_xlsx_empty_export_is_valid(
    client: AsyncClient, email_backend: RecordingEmailBackend
) -> None:
    await login_user_a(client, email_backend)
    response = await client.get("/api/v1/export/expenses", params={"format": "xlsx"})
    assert response.status_code == 200
    _, rows = _load_sheet(response.content)
    assert rows == [("Date", "Day", "Category", "Description", "Amount")]
